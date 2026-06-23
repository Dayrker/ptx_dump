#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成「启动 NCCL device PTX 所需的 CPU/driver/host 支持」依赖表 (Excel)。

重点：device 函数我们做 PTX 兼容；本表回答「配套的 driver/cpu/host 函数要怎么支持」——
即把一份 ncclDevKernel_AllReduce_..._RING_LL.ptx 真正启动起来，软件团队要补哪些能力。

唯一真源：docs/allreduce-deep-dive.md + nccl_ptx/call_chains.json / CALL_CHAINS.txt 实测 trace。
实测锚点（双卡 Qwen3-8B）：
  f16 all_reduce : RING+LL, grid=[1,1,1], block=[512,1,1], smem=88416, 4608 次
  f32 barrier    : RING+LL, grid=[1,1,1], block=[96,1,1],  smem=88416, 1 次
两条链路 host 选择/smem/launch 形态完全相同，只有 count/block 随数据量变。
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

wb = Workbook()

# 列含义（主表）：
#  阶段       = 冷启动建链 / 每次通信热路径 / driver 装载启动 / device 翻译
#  能力/函数  = 具体要支持的 CPU/driver/host 函数或能力
#  归属团队   = 软件(host/runtime) / 驱动(driver) / 硬件 / 翻译(device PTX)
#  作用       = 在启动这份 PTX 的流程里干什么
#  不做会怎样 = 缺了它 PTX 启动会卡在哪
#  必选/可选  = 必选 / 可选 / 复用NCCL则白送
#  难度       = ★ 低 / ★★ 中 / ★★★ 高(易卡点)

# ============================================================
# Sheet 0: 先读这页（导图）
# ============================================================
ws = wb.active
ws.title = "先读这页"
ws.append(["项", "说明", "对应"])
guide_rows = [
    ("这张表回答什么",
     "device 函数我们能做 PTX 兼容；但 device kernel 不自己启动。本表列出「让这份 PTX 跑起来，软件/驱动团队要补的 CPU/driver/host 函数」。",
     "allreduce-deep-dive.md 正文"),
    ("一句话主线",
     "PyTorch 交参数 → NCCL host 选算法/造任务单 → driver 装 PTX+P2P+launch → device kernel 才执行。前三步全是 CPU/driver 侧，是本表重点。",
     "onepager.md / deep-dive §0"),
    ("怎么用这个工作簿",
     "①先看本页；②主看「软件团队任务清单」决定要实现哪些 host/driver 能力；③「driver-runtime原语」「device层ISA要求」给具体团队;④「实测事实卡片」是所有数字的唯一真源；⑤「验证检查清单」按 P0/P1/P2 验收。",
     "本工作簿各 sheet"),
    ("最重要的三件事(易卡点)",
     "① 跨卡 P2P 内存映射(Ring LL 的 ld/st.volatile.global 全靠它)；② 88416B 大 shared mem 需 opt-in(>48KB)；③ 要 load 整个 device 模块(PTX 引用未声明的 ncclShmem/ncclDevFuncTable)。",
     "软件团队任务清单 难度★★★ 行"),
    ("两条落地策略",
     "策略1=复用 NCCL host，软件只补 driver/runtime+翻译 PTX(工作量最小，推荐先走)；策略2=host 也自研 drop-in 兼容库(工作量大)。两条路的硬约束相同。",
     "落地策略与约束"),
    ("barrier 不是额外工作量",
     "NCCL 无独立 barrier；dist.barrier 复用 AllReduce(1 元素)。打通 RING+LL 即同时覆盖 all_reduce 和 barrier。",
     "deep-dive §1 / 附录A"),
    ("⚠ 已纠正旧版错误",
     "旧文档写 f16 grid=[16,1,1]/channelMask=0xFFFF/volatile=18/全inline——均与实测不符。实测 grid=[1,1,1]、只 1 个 channel、f16 有 21 条 volatile、PTX 仍引用外部符号+1 条间接 call。",
     "实测事实卡片"),
]
for r in guide_rows:
    ws.append(list(r))

# ============================================================
# Sheet 1: 软件团队任务清单 (主表)
# ============================================================
ws1 = wb.create_sheet("软件团队任务清单")
headers = ["阶段", "能力 / 函数", "归属团队", "作用(在启动 PTX 里干什么)",
           "不做会怎样", "必选/可选", "难度"]
ws1.append(headers)

rows = [
    # ── 阶段 A：冷启动 / 建链（communicator 生命周期一次性）──
    ("A 冷启动建链", "ncclGetUniqueId / ncclCommInitRank(All)", "软件(host) / 复用NCCL则白送",
     "建 communicator：定 rank/nRanks、算 RING 拓扑(prev/next)、开 ring buffer、在 device memory 摆好 ncclDevComm/channels",
     "没有 comm，ncclAllReduce 无从调用；device kernel 拿不到通信上下文", "必选", "★★"),
    ("A 冷启动建链", "跨卡 P2P 内存映射 cudaDeviceEnablePeerAccess + IPC 句柄", "驱动 + 硬件",
     "让本卡能直接读写对端 GPU 暴露的 ring buffer(NVLink/PCIe P2P)",
     "Ring LL 的 ld/st.volatile.global 跨卡访问失败，通信根本跑不起来", "必选(关键)", "★★★"),
    ("A 冷启动建链", "cuModuleLoadData / cuModuleGetFunction", "驱动",
     "把整个 device 模块(含 ncclShmem/ncclDevFuncTable)装进 GPU、取 kernel 句柄",
     "PTX 装不进去，连函数指针都没有", "必选", "★★"),
    ("A 冷启动建链", "cuMemAlloc / cuMemcpyHtoD (cudaMalloc/cudaMemcpy)", "驱动",
     "为 devComm/channel/ring buffer/abortFlag 分配显存并填充",
     "device 端结构体无处安放", "必选", "★"),

    # ── 阶段 B：每次 all_reduce 热路径 ──
    ("B 每次通信", "ncclAllReduce(send,recv,count,dtype,op,comm,stream)", "软件(host 公共API) / 复用NCCL则白送",
     "NCCL C API 入口；构造 ncclInfo 后转入 enqueue。PyTorch 直接调它",
     "PyTorch 的 c10d 路径接不上，all_reduce 报错", "必选", "★"),
    ("B 每次通信", "枚举 ncclDataType_t / ncclRedOp_t / ncclComm_t", "软件(host 公共API)",
     "torch 侧 getNcclDataType/getNcclReduceOp 依赖其数值；f16→ncclFloat16、SUM→ncclSum",
     "类型/归约枚举对不上，参数语义错乱", "必选", "★"),
    ("B 每次通信", "ncclEnqueueCheck → taskAppend (+hostToDevRedOp)", "软件(host 内部) / 复用NCCL则白送",
     "校验 comm/参数；ncclSum→ncclDevSum；单rank memcpy快速路径；多rank入 collQueue",
     "任务进不了队，后续调度无输入", "必选*", "★★"),
    ("B 每次通信", "scheduleCollTasksToPlan (+topoGetAlgoInfo/ncclDevFuncId/getChannnelThreadInfo/computeCollChunkInfo/initCollWorkElem)", "软件(host 内部) / 复用NCCL则白送",
     "★选 RING+LL、算 channel/线程(实测 grid=1/block=512)、定 funcIndex、填 ncclWork 任务单",
     "选不出算法、产不出 PTX 能读懂的任务单，kernel 行为未定义", "必选*", "★★★"),
    ("B 每次通信", "uploadWork → workFifo", "软件(host 内部) / 复用NCCL则白送",
     "把 ncclWork 拷到 device 的 workFifo，得到 workHead 指针(launch 第3参数)",
     "device kernel 主循环读不到 work，空转/崩溃", "必选*", "★★"),
    ("B 每次通信", "ncclGroupStart / ncclGroupEnd", "软件(host) / 复用NCCL则白送",
     "group 批提交语义(depth==1 触发实际 launch)",
     "group 语义缺失，提交时机错乱", "必选*", "★"),

    # ── 阶段 C：driver 装载 + 启动 ──
    ("C driver启动", "cudaFuncSetAttribute(MaxDynamicSharedMemorySize=88416)", "驱动 + 硬件",
     "申请 88416B 动态 shared mem(>48KB 必须 opt-in)，硬件 shared 容量要够",
     "smem>48KB 不 opt-in → launch 直接失败", "必选(关键)", "★★★"),
    ("C driver启动", "cudaLaunchKernelExC / cuLaunchKernel (+cudaLaunchConfig_t)", "驱动",
     "按 grid/block/dynamicSmemBytes/stream + 3 个参数{&devComm,&channelMask,&workHead} 启动",
     "kernel 起不来；或参数布局错位", "必选", "★★"),
    ("C driver启动", "cudaStream_t + cudaEvent + syncStream", "驱动",
     "NCCL 用独立 comms stream，与计算 stream 异步重叠+同步",
     "通信/计算无法重叠或乱序，正确性/性能均受损", "必选", "★★"),
    ("C driver启动", "cuCtxSynchronize / stream sync", "驱动",
     "等 kernel 完成、收尾",
     "无法判定通信完成", "必选", "★"),

    # ── 阶段 D：device 翻译（PTX 兼容本体，列出依赖以闭环）──
    ("D device翻译", "解析外部符号 ncclShmem(静态shared) / ncclDevFuncTable(device global)", "翻译(device) + 驱动",
     "PTX 引用但不声明这两个符号；必须 load 整个 NCCL device 模块、能解析它们",
     "单独 load 一个 entry → 符号未定义，装载失败", "必选(关键)", "★★★"),
    ("D device翻译", "ld/st.volatile.global(含 .v4.u32) 跨卡语义", "翻译(device) + 驱动",
     "LL 协议 data+flag 打包，跨卡读写 peer ring buffer(实测 f16 共 21 条)",
     "通信数据搬不动，结果错", "必选(关键)", "★★★"),
    ("D device翻译", "bar.sync 0 / bar.sync %r,%n / bar.warp.sync / bar.red.popc", "翻译(device)",
     "全block/指定线程数/ warp 内同步 + 带规约的 barrier(线程数非固定:512/96)",
     "同步错乱 → 数据竞争/死锁", "必选", "★★"),
    ("D device翻译", "popc.b64 / cvta.shared / shf.r.wrap / bfi 等", "翻译(device)",
     "channelMask→channelId 映射、地址空间转换、LL 打包/位运算",
     "地址/位图计算错，访存越界", "必选", "★★"),
    ("D device翻译", "device 端间接 call(经 ncclDevFuncTable)", "翻译(device)",
     "非特化路径(LL128/SIMPLE/Generic)经函数表分发；本 PTX 保留 1 条",
     "兜底分发路径不可用(LL 特化路径可暂时绕过)", "可选(本例可暂缓)", "★★"),

    # ── 仅 trace/profiling，非功能路径 ──
    ("(可选)非功能", "record_param_comms", "软件(host)",
     "param↔comms 映射，仅 profiling/DTensor 用，非启动 PTX 必需",
     "不影响功能，仅 trace 缺信息", "可选", "★"),
]
for r in rows:
    ws1.append(list(r))

# ============================================================
# Sheet 2: driver / runtime 原语清单（给驱动团队）
# ============================================================
ws2 = wb.create_sheet("driver-runtime原语")
ws2.append(["原语", "用在哪一步", "为什么不可省", "难度"])
drv_rows = [
    ("cuModuleLoadData / cuModuleGetFunction", "阶段A 装 PTX 模块", "拿不到 kernel 句柄就无法 launch", "★★"),
    ("cuMemAlloc / cuMemcpyHtoD", "摆 devComm/channel/work/buffer", "device 结构体需落显存", "★"),
    ("cudaDeviceEnablePeerAccess + IPC", "阶段A 建跨卡 ring buffer 互访", "Ring LL 跨卡 volatile 读写全靠它(最易卡)", "★★★"),
    ("cudaFuncSetAttribute(MaxDynamicSharedMemorySize)", "launch 前 opt-in 88416B", ">48KB 不 opt-in 直接 launch 失败", "★★★"),
    ("cudaLaunchKernelExC / cuLaunchKernel", "阶段C launch", "需支持 cudaLaunchConfig_t(grid/block/dynSmem/stream) + cluster/调度 attr", "★★"),
    ("cudaStream_t + cudaEvent + 同步", "全程", "独立 comms stream 与计算异步重叠", "★★"),
    ("cuCtxSynchronize / stream sync", "收尾", "判定通信完成", "★"),
    ("大 shared mem 硬件容量(sm 级)", "硬件前提", "88416B/CTA 需硬件 shared 容量支撑", "★★★"),
]
for r in drv_rows:
    ws2.append(list(r))

# ============================================================
# Sheet 3: device 层 ISA 要求（给做 PTX 翻译的团队）
# ============================================================
ws3 = wb.create_sheet("device层ISA要求")
ws3.append(["项", "内容(实测)", "出处", "说明 / 兼容要点"])
dev_rows = [
    ("翻译目标 entry",
     ".visible .entry ncclDevKernel_AllReduce_Sum_{f16,f32}_RING_LL",
     "nccl_ptx/*.ptx",
     "已 inline ncclKernelMain/RunWork/runRing/ProtoLL；但仍引用外部符号+1 条间接 call(非全 inline)"),
    ("外部符号 ncclShmem",
     "静态 shared 块，按固定偏移访问(本 kernel 到 +5655B)",
     "deep-dive §3.1/4",
     "host 摆的 ncclDevComm/work 拷进它；layout 必须与 PTX 偏移一致(ABI)"),
    ("外部符号 ncclDevFuncTable",
     "device global 函数表，间接 call 目标",
     "deep-dive §3.1",
     "必须 load 整个模块才能解析；不能只 load 单 entry"),
    ("ISA: ld/st.volatile.global(.v4.u32)",
     "f16 共 21 条 (f32 为 18)",
     "PTX 实测",
     "跨卡 P2P 读写 peer ring buffer；LL data+flag；通信核心"),
    ("ISA: bar.sync 0 / bar.sync %r,%n",
     "全block / 指定线程数同步",
     "PTX 实测",
     "线程数非固定(512/96)，NCCL warp 分工"),
    ("ISA: bar.warp.sync / bar.red.popc.u32",
     "warp 内同步 / barrier+popcount 规约",
     "PTX 实测",
     "带规约语义的 barrier 需正确翻译"),
    ("ISA: popc.b64",
     "channelMask(64bit位图)→channelId",
     "PTX 实测",
     "__popcll；grid.x=popcount(channelMask)，实测=1"),
    ("ISA: cvta.shared / cvta.to.global",
     "地址空间转换(5 条)",
     "PTX 实测",
     "shared/global 地址空间模型要一致"),
    ("ISA: shf.r.wrap.b32 / bfi.b64",
     "funnel-shift / 位域插入",
     "PTX 实测",
     "LL 打包、地址/位计算"),
    ("ISA: 间接 call",
     "经 ncclDevFuncTable[funcIndex]",
     "PTX 实测(line 1517-1522)",
     "需支持 device 端间接调用；LL 特化路径可暂绕过"),
    ("dynamic smem",
     "88416 B(协议无关取 max)",
     "call_chains.json",
     "launch 参数，非 0；区别于静态 ncclShmem(~5.7KB)"),
]
for r in dev_rows:
    ws3.append(list(r))

# ============================================================
# Sheet 4: 实测事实卡片（唯一真源）
# ============================================================
ws4 = wb.create_sheet("实测事实卡片")
ws4.append(["项", "f16 AllReduce (主线)", "f32 AllReduce (barrier 复用)", "备注"])
fact_rows = [
    ("PTX 文件", "nccl_001_..._f16_RING_LL_...ptx", "nccl_002_..._f32_RING_LL_...ptx", "dump 自本地 libnccl.so"),
    ("触发入口", "dist.all_reduce() (TP MLP/Attn 聚合)", "dist.barrier() (同步)", "barrier 复用 AllReduce 机器"),
    ("algo × proto", "RING + LL", "RING + LL", "topoGetAlgoInfo 暴力选最小时间"),
    ("grid", "[1,1,1]", "[1,1,1]", "★只 1 个 channel；旧文档写[16,1,1]是错的"),
    ("block", "[512,1,1]", "[96,1,1]", "LL≤512，按数据量调谐最低 96"),
    ("dynamic smem", "88416 B", "88416 B", "两条链路一致；>48KB 需 opt-in"),
    ("count(元素数)", "4096 (也见 16×4096)", "1", "count 是元素数不是字节"),
    ("launch API", "cudaLaunchKernelExC", "cudaLaunchKernelExC", "CUDA 11.8+ 路径"),
    ("调用次数/累计", "4608 次 / 58.06 ms", "1 次 / 0.01 ms", "f16 是真正热点"),
    ("volatile.global 条数", "21", "18", "旧文档误把 18 当 f16 的数"),
    ("外部符号", "ncclShmem + ncclDevFuncTable", "同左", "PTX 引用但不声明，需 load 整模块"),
    ("结论", "host 选择/smem/launch 形态两条完全相同，只 count/block 变", "", "打通 RING+LL 即同时覆盖两条"),
]
for r in fact_rows:
    ws4.append(list(r))

# ============================================================
# Sheet 5: 落地策略与约束
# ============================================================
ws5 = wb.create_sheet("落地策略与约束")
ws5.append(["项", "内容"])
strat_rows = [
    ("策略1: 复用 NCCL host + 换 device(推荐先走)",
     "NCCL host 全留(任务清单里标*的全部白送)，软件团队只补阶段C driver/runtime(load/P2P/大smem/launch/stream)和阶段D 翻译 device PTX。"
     "工作量最小；前提:runtime 能被 NCCL host 当 CUDA runtime 用、driver API 兼容度高。"),
    ("策略2: host 也自研(drop-in 兼容库)",
     "自研 ncclAllReduce+comm 初始化+枚举(公共 API)，内部自造 ncclWork。标*的 host 内部函数不必逐个照抄函数名，"
     "但必须最终产出与 PTX 期望一致的 devComm/ncclWork/channelMask 布局。工作量大，ABI 风险高。"),
    ("跨层硬约束1: ABI",
     "host 填进 device memory 的 ncclDevComm/ncclDevChannel/ncclWork/ncclWorkElem 字段偏移，必须和 device PTX 里 ld.shared 的偏移严丝合缝。"),
    ("跨层硬约束2: 整模块装载",
     "PTX 引用未声明的 ncclShmem/ncclDevFuncTable，必须 load 整个 NCCL device 模块(cuobjdump -ptx 的完整模块)，不能只 load 单 entry。"),
    ("跨层硬约束3: P2P",
     "Ring LL 的跨卡 ld/st.volatile.global 依赖跨卡 P2P 内存映射，硬件/驱动必须支持(最易卡点)。"),
    ("跨层硬约束4: 大 shared mem",
     "smem=88416>48KB，必须 cudaFuncSetAttribute opt-in，且硬件 shared 容量要够。"),
    ("barrier 复用",
     "barrier 与 all_reduce 共用阶段B/C/D，只入口 aten op 不同(c10d::barrier vs c10d::allreduce_)；非额外工作量。"),
    ("图例",
     "必选=启动 PTX 必需；必选*=Drop-in 策略下需等价实现，复用 NCCL host 策略下白送；可选=仅 trace/profiling 或本例可暂缓。"
     "难度:★低/★★中/★★★高(易卡点)。"),
]
for r in strat_rows:
    ws5.append(list(r))

# ============================================================
# Sheet 6: 验证检查清单
# ============================================================
ws6 = wb.create_sheet("验证检查清单")
ws6.append(["优先级", "检查项", "通过标准", "对应依据"])
check_rows = [
    ("P0", "整 device 模块可装载",
     "cuModuleLoadData 能 load 含 ncclShmem/ncclDevFuncTable 的完整模块并取到 kernel 句柄(非单 entry)。",
     "device层ISA要求 / deep-dive §3"),
    ("P0", "跨卡 P2P 可用",
     "peer GPU ring buffer 能被本卡 ld/st.volatile.global 访问，volatile 语义正确。",
     "driver-runtime原语 / 任务清单 A 行"),
    ("P0", "大 shared mem opt-in",
     "cudaFuncSetAttribute(MaxDynamicSharedMemorySize=88416) 成功，且能以 smem=88416 launch。",
     "实测事实卡片 / 任务清单 C 行"),
    ("P0", "device 结构体 ABI 兼容",
     "ncclDevComm/ncclDevChannel/ncclWork/ncclWorkElem 字段偏移与 PTX 的 ld.shared 偏移一致。",
     "落地策略 硬约束1 / deep-dive 附录C"),
    ("P0", "launch 三参数 + 形态一致",
     "传 {&devComm,&channelMask,&workHead}；grid.x=popcount(channelMask)；实测 grid=1。",
     "实测事实卡片 / deep-dive §4"),
    ("P0", "AllReduce 参数语义",
     "count 按元素数解释；sendbuff/recvbuff 支持 in-place；stream 异步语义保持。",
     "deep-dive 附录B.2"),
    ("P1", "RING+LL 主线可跑通",
     "复现 f16: grid=1/block=512/smem=88416；f32 barrier: grid=1/block=96。",
     "实测事实卡片"),
    ("P1", "barrier 复用 AllReduce",
     "torch barrier 走 count=1 的 ncclAllReduce，不虚构 ncclBarrier。",
     "deep-dive 附录A"),
    ("P1", "ISA 覆盖",
     "支持 ld/st.volatile.global、bar.sync(含%r,%n)/bar.warp.sync/bar.red.popc、popc、cvta、shf/bfi。",
     "device层ISA要求"),
    ("P2", "Generic/间接 call 边界",
     "LL128/SIMPLE 不误用 RING_LL 特化 entry；需要时经 ncclDevFuncTable 走 Generic。",
     "deep-dive 附录B.4 / §3.1"),
]
for r in check_rows:
    ws6.append(list(r))

# ============================================================
# Sheet 7: 术语表
# ============================================================
ws7 = wb.create_sheet("术语表")
ws7.append(["术语", "一句话理解", "为什么重要"])
term_rows = [
    ("AllReduce", "每卡各有一份数据，调用后每卡都拿到所有 rank 归约后的完整结果。", "TP 推理里聚合 MLP/Attn 输出，是本链路的 collective。"),
    ("rank", "通信组里的 GPU 编号；双卡是 rank 0/1。", "ring 拓扑、prev/next、buffer 都按 rank 组织。"),
    ("communicator / ncclComm", "一组 GPU 的通信上下文(rank 数/拓扑/channel/device 结构体地址)。", "每次 all_reduce 都要先有 comm 才能调 ncclAllReduce。"),
    ("channel", "把一次通信拆成的并行车道。", "实测本次只用 1 个 channel，故 grid.x=1。"),
    ("algo / proto", "拓扑算法(RING/TREE/NVLS) / 传输协议(LL/LL128/SIMPLE)。", "实测选中 RING+LL；本地特化 kernel 都是 LL。"),
    ("ncclWork / ncclWorkElem", "host 写给 device kernel 的任务单(send/recv buffer、count、chunk、funcIndex)。", "自研 host 不必照抄函数名，但任务单布局必须 PTX 能读懂。"),
    ("channelMask", "64bit 位图，哪 bit 为 1 即哪 channel 参与。", "kernel 用 __popcll 把 blockIdx→channelId；grid.x=popcount，实测=1。"),
    ("ncclShmem", "device 端静态 shared 块，kernel 按固定偏移读写。", "host 摆的结构体拷进它；layout 必须与 PTX 偏移一致(ABI)。"),
    ("ncclDevFuncTable", "device global 函数表，间接 call 目标。", "PTX 引用但不声明，必须 load 整模块。"),
    ("P2P memory mapping", "本卡直接访问对端 GPU 暴露的 ring buffer。", "Ring LL 的 ld/st.volatile.global 依赖它(最易卡)。"),
    ("dynamic shared memory", "launch 时额外申请的 shared 字节数(实测 88416)。", ">48KB 必须 opt-in；不能以为 LL 就是 0。"),
    ("drop-in 兼容库", "实现同名同签名 NCCL API、可被 torch 直接链接的库。", "策略2 的形态；策略1 则复用真 NCCL host。"),
]
for r in term_rows:
    ws7.append(list(r))

# ============================================================
# 样式
# ============================================================
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")

# 阶段 / 团队配色
PHASE_FILLS = {
    "A ": PatternFill("solid", fgColor="DDEBF7"),   # 蓝 - 建链
    "B ": PatternFill("solid", fgColor="FFF2CC"),   # 黄 - 热路径
    "C ": PatternFill("solid", fgColor="FCE4D6"),   # 橙 - driver
    "D ": PatternFill("solid", fgColor="E2F0D9"),   # 绿 - device
    "(可": PatternFill("solid", fgColor="F2F2F2"),  # 灰 - 可选
}
PRIORITY_FILLS = {
    "P0": PatternFill("solid", fgColor="F4CCCC"),
    "P1": PatternFill("solid", fgColor="FCE5CD"),
    "P2": PatternFill("solid", fgColor="D9EAD3"),
}

def style_sheet(ws, widths, color_first=True):
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP
            c.border = BORDER
        if not color_first:
            continue
        first = str(row[0].value or "")
        for key, fill in PHASE_FILLS.items():
            if first.startswith(key):
                row[0].fill = fill
                break
        if first in PRIORITY_FILLS:
            row[0].fill = PRIORITY_FILLS[first]
            row[0].font = Font(bold=True)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 46

style_sheet(wb["先读这页"],       [22, 92, 30], color_first=False)
style_sheet(ws1, [16, 40, 22, 50, 40, 14, 8])
style_sheet(ws2, [40, 26, 50, 8])
style_sheet(ws3, [26, 40, 22, 50])
style_sheet(ws4, [20, 40, 40, 40], color_first=False)
style_sheet(ws5, [30, 100], color_first=False)
style_sheet(ws6, [10, 30, 60, 36])
style_sheet(ws7, [24, 56, 60], color_first=False)

out = Path(__file__).with_name("ptx-compat-requirements.xlsx")
wb.save(out)
print("written:", out)
